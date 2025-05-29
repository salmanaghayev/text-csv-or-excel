import re
import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
import os
import os.path

# Define pattern-replacement rules
replacements = [
    (r'\s{2,}', ';'),     # Replace 2+ spaces with semicolon
    (r'\t+', ';'),        # Replace tabs with semicolon
    (r'\s+\|\s+', '|'),   # Replace space-pipe-space with pipe
    (r':', '=')           # Replace colon with equals
]

def process_line(line):
    line = line.strip('"')  # Remove leading/trailing quotes
    for pattern, replacement in replacements:
        line = re.sub(pattern, replacement, line)
    return line

def main(input_files, additional_files, excel_file):
    # Load or create workbook
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    for input_file, additional_file in zip(input_files, additional_files):
        sheet_name = os.path.splitext(os.path.basename(input_file))[0]

        # Read and clean lines from input file
        with open(input_file, 'r', encoding='utf-8') as infile:
            lines = [process_line(line) for line in infile]

        rows = [line.split(';') for line in lines]

        # Read corresponding additional file and map third col to a list of first col values
        additional_map = {}
        if os.path.exists(additional_file):
            with open(additional_file, 'r', encoding='utf-8') as adfile:
                for line in adfile:
                    parts = process_line(line).split(';')
                    if len(parts) >= 3:
                        key = parts[2].strip()
                        value = parts[0].strip()
                        additional_map.setdefault(key, []).append(value)

        # Add mapped value to each row from input file where input[0] matches additional[2]
        for i, row in enumerate(rows):
            if i == 0:
                row.append("Extra Info")
            else:
                input_key = row[0].strip()
                extra_values = additional_map.get(input_key, [])
                row.append(", ".join(extra_values))

        # Delete existing sheet if present
        if sheet_name in wb.sheetnames:
            print(f"Sheet '{sheet_name}' already exists. Overwriting.")
            del wb[sheet_name]

        # Create new sheet
        ws = wb.create_sheet(title=sheet_name)

        # Define header style
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=12)
        header_alignment = Alignment(vertical="center", horizontal="center")

        # Define thick border style
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

        # Write data to the new sheet
        col_widths = dict()
        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                value = value.strip()
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                col_letter = get_column_letter(col_idx)
                col_widths[col_letter] = max(col_widths.get(col_letter, 0), len(value) + 2)
                cell.border = thick_border
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

        # Adjust column widths
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = ws['A2']

        # Add filter to the headers
        ws.auto_filter.ref = ws.dimensions

        # Add conditional formatting for entire row based on columns B and C
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        green_dxf = DifferentialStyle(fill=green_fill)
        red_dxf = DifferentialStyle(fill=red_fill)

        max_row = ws.max_row
        max_col = ws.max_column
        last_col_letter = get_column_letter(max_col)

        formula_green = "=AND(EXACT($B2,\"up\"), EXACT($C2,\"up\"))"
        formula_red = "=NOT(AND(EXACT($B2,\"up\"), EXACT($C2,\"up\")))"

        ws.conditional_formatting.add(f"A2:{last_col_letter}{max_row}", Rule(type="expression", formula=[formula_green], dxf=green_dxf))
        ws.conditional_formatting.add(f"A2:{last_col_letter}{max_row}", Rule(type="expression", formula=[formula_red], dxf=red_dxf))

    # Save workbook
    try:
        wb.save(excel_file)
        print(f"Data written to '{excel_file}'.")
    except PermissionError as e:
        print(f"❌ Permission error saving file: {e}")
    except Exception as e:
        print(f"❌ Unexpected error: {e}")

if __name__ == "__main__":
    input_files = [
        "input1.txt",
        "input2.txt"
    ]
    additional_files = [
        "additional1.txt",
        "additional2.txt"
    ]
    excel_file = "output.xlsx"
    main(input_files, additional_files, excel_file)
